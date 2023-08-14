from PIL import Image, ImageDraw, ImageFont
import pandas as pd
import cv2
import shutil
import os 
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def get_info(xlsx_filename):
    nested_dict = defaultdict(dict)
    workbook = openpyxl.load_workbook(xlsx_filename)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        picture = row[0]
        lane = int(row[1])
        xulab_id = row[2]
        if picture not in nested_dict:
            nested_dict[picture]['lane_num'] = lane + 1
        else:
            nested_dict[picture]['lane_num'] = max(nested_dict[picture]['lane_num'], lane + 1)
        nested_dict[picture][f'lane{lane}'] = xulab_id
    return nested_dict

def cut(result_dict, picture, marker_lane=0, center=0, move=0):
    # 打开图片
    image_path = f"{picture}.jpg"
    img = Image.open(image_path)
    # img = img.convert("RGB")
    
    # 获取图片宽度和高度
    width, height = img.size
    
    # 计算每个切割部分的宽度
    num_parts = result_dict[f'{picture}']["lane_num"] # 切割成的部分数目
    part_width = width // num_parts
    
    # 切割并保存图片
    for i in range(num_parts):
        left = max(0, (i + center + move) * part_width)
        right = min(width, (i + 1 - center + move) * part_width if i < num_parts - 1 else width)

        # 根据切割位置获取子图
        part = img.crop((left, 0, right, height))
        
        # 设置保存文件名
        if i == marker_lane:
            filename = os.path.join(picture, 'marker.jpg')
            part.save(filename)
        else:
            filename = os.path.join(picture, f'lane{i}.jpg')
            part.save(filename)

def marker(marker_thresh,marker_list):
    for pic,thresh in marker_thresh.items():
        marker_path=f"{pic}/marker.jpg"
        imgBGR = cv2.imread(marker_path)
        img = cv2.cvtColor(imgBGR, cv2.COLOR_BGR2RGB)
        gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        thresh, binary = cv2.threshold(gray, thresh, 255, cv2.THRESH_BINARY)
        contours, hierarchy = cv2.findContours(~binary, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE) #binary反转是为了不把最外层的轮廓框出来

        contours = sorted(contours, key=lambda c: cv2.boundingRect(c)[2], reverse=True)
        max_width = cv2.boundingRect(contours[0])[2]
        contours = [c for c in contours if cv2.boundingRect(c)[2] > max_width * 0.2]

        rectangles = [cv2.boundingRect(cnt) for cnt in contours]
        rectangles.sort(key=lambda x: x[1])  # 按照 y 坐标进行排序
        merged_rectangles = []
        current_rect = rectangles[0]

        for rect in rectangles[1:]:
            if rect[1] - current_rect[1] <= current_rect[3]:  # 判断是否有重叠
                current_rect = (
                    min(current_rect[0], rect[0]),
                    min(current_rect[1], rect[1]),
                    max(current_rect[0] + current_rect[2], rect[0] + rect[2]) - min(current_rect[0], rect[0]),
                    max(current_rect[3], rect[3])
                )
            else:
                merged_rectangles.append(current_rect)
                current_rect = rect
        merged_rectangles.append(current_rect)

        img_contours = img.copy()
        # for rect in merged_rectangles:
        #     x, y, w, h = rect
        #     cv2.rectangle(img_contours, (x, y), (x + w, y + h), (255, 0, 0), 2)

        img_contours = Image.fromarray(img_contours)

        image = Image.new('RGBA', (img_contours.width + 40, img_contours.height))
        image.paste(img_contours, (40, 0))


        merged_rectangles = sorted(merged_rectangles, key=lambda c: c[1], reverse=True)
        for i, rect in enumerate(merged_rectangles):
            x, y, w, h = rect
            marker_text = str(marker_list[i])
            font = ImageFont.truetype("/Library/Fonts/Arial Unicode.ttf", 20)
            draw = ImageDraw.Draw(image)
            marker_width, marker_height = draw.textsize(marker_text, font=font)
            position = (35-marker_width, y-6+0.5*h-0.5*marker_height)  
            draw.text(position, marker_text, font=font, fill=(0, 0, 0, 255))  # 黑色文字

        image.save(f'{marker_path[:-4]}.png', format='PNG')

def concatenate(img1, img2, title, output_path):
    # 打开两张图片
    img1 = Image.open(img1)
    img1 = img1.crop((0, 0, 40, img1.height))
    img2 = Image.open(img2)
    # 将图片1和图片2左右拼接
    cat_img = Image.new('RGBA', (img1.width + img2.width, img1.height + 40))  # 增加高度以容纳标题
    cat_img.paste(img1, (0, 40))
    cat_img.paste(img2, (img1.width, 40))

    font = ImageFont.truetype("/Library/Fonts/Arial Unicode.ttf", 20)
    draw = ImageDraw.Draw(cat_img)

    title_width, title_height = draw.textsize(title, font=font)
    title_position = (40 + (img2.width-title_width) // 2,(40 - title_height) // 2)
    draw.text(title_position, title, font=font, fill=(0, 0, 0, 255))  # 黑色文字

    marker_title = "kDa"
    marker_title_width, marker_title_height = draw.textsize(marker_title, font=font)
    marker_title_position = (35 - marker_title_width, (40 - marker_title_height))
    draw.text(marker_title_position, marker_title, font=font, fill=(0, 0, 0, 255))  # 黑色文字

    # 保存拼接后的图片
    cat_img.save(output_path, format='PNG')

def add_hyperlinker(xlsx_filename, picture, target_text, local_file_path):
    workbook = openpyxl.load_workbook(xlsx_filename)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2):  # 从第2行开始遍历，第一行为表头
        if row[0].value == picture:
            for cell in row:
                if cell.value == target_text:
                    cell.hyperlink = local_file_path
                    cell.style = 'Hyperlink'
                elif cell.value == "M":
                    cell.hyperlink = f"{picture}.jpg"
                    cell.style = 'Hyperlink'
    new_file_path = xlsx_filename
    workbook.save(new_file_path)
